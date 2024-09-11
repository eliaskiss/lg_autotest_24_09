import sys

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side
from icecream import ic

from lec_pymysql import Database

# Talbe Name : 자신의계정_bicycle

# Create Table SQL
# table_name 지정필요!!!!
# CREATE TABLE elias_bicycle (
#   `id` INT(11) NOT NULL AUTO_INCREMENT,
#   `reg_datetime` DATETIME DEFAULT CURRENT_TIMESTAMP(),
#   `station_number` INT(11) DEFAULT NULL,      # 대여소 번호
#   `station_name` VARCHAR(128) DEFAULT NULL,   # 보관소(대영소) 명
#   `region` VARCHAR(128) DEFAULT NULL,         # 자치구
#   `address` VARCHAR(1024) DEFAULT NULL,       # 상세주소
#   `latitude` FLOAT DEFAULT NULL,              # 위도
#   `longitude` FLOAT DEFAULT NULL,             # 경도
#   `install_date` DATETIME DEFAULT NULL,       # 설치시기
#   `lcd_count` INT(11) DEFAULT NULL,           # LCD 거치대수
#   `qr_count` INT(11) DEFAULT NULL,            # QR 거치대수
#   `proc_type` VARCHAR(128) DEFAULT NULL,      # 운영방식
#   KEY `id` (`id`)
# ) ENGINE=INNODB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci;

# 'CREATE TABLE if not exists elias_bicycle ' \
# '(`id` INT(11) NOT NULL AUTO_INCREMENT, ' \
# '`reg_datetime` DATETIME DEFAULT CURRENT_TIMESTAMP(), ' \
# '`station_number` INT(11) DEFAULT NULL, ' \
# '`station_name` VARCHAR(128) DEFAULT NULL, ' \
# '`region` VARCHAR(128) DEFAULT NULL, ' \
# '`address` VARCHAR(1024) DEFAULT NULL, ' \
# '`latitude` FLOAT DEFAULT NULL, ' \
# '`longitude` FLOAT DEFAULT NULL, ' \
# '`install_date` DATETIME DEFAULT NULL, ' \
# '`lcd_count` INT(11) DEFAULT NULL, ' \
# '`qr_count` INT(11) DEFAULT NULL, ' \
# '`proc_type` VARCHAR(128) DEFAULT NULL, ' \
# 'KEY `id` (`id`)' \
# ') ENGINE=INNODB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci;'

# CREATE TABLE IF NOT EXISTS 자신의계정_bicycle
# (`id` INT(11) NOT NULL AUTO_INCREMENT,
# `reg_datetime` DATETIME DEFAULT CURRENT_TIMESTAMP(),
# `station_number` INT(11) DEFAULT NULL,
# `station_name` VARCHAR(128) DEFAULT NULL,
# `region` VARCHAR(128) DEFAULT NULL,
# `address` VARCHAR(1024) DEFAULT NULL,
# `latitude` FLOAT DEFAULT NULL,
# `longitude` FLOAT DEFAULT NULL,
# `install_date` DATETIME DEFAULT NULL,
# `lcd_count` INT(11) DEFAULT NULL,
# `qr_count` INT(11) DEFAULT NULL,
# `proc_type` VARCHAR(128) DEFAULT NULL,
# KEY `id` (`id`)
# ) ENGINE=INNODB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci;

DB_URL = '139.150.73.242'
DB_USER = 'dbuser'
DB_PW = 'dbuser'
DB_NAME = 'lg_autotest'

table_name = '자신의계정_bicycle'


###################################################################################################
# Task 1
# public_bicycle.xlsx 파일을 읽어서, DB의 자신의계정_bicycle의 테이블에 넣기 (ex: elias_bicycle)
###################################################################################################
def put_data_to_db(excel_file_name):
    # Load wb from excel file
    wb = load_workbook(excel_file_name, data_only=True, read_only=True)

    # Select work sheet
    ws = wb['대여소현황']

    # DB 객체생성
    db = Database(DB_URL, DB_USER, DB_PW, DB_NAME)

    # DB 연결
    db.connect_db()

    # Table 생성
    sql = f'CREATE TABLE if not exists {table_name} ' \
          '(`id` INT(11) NOT NULL AUTO_INCREMENT, ' \
          '`reg_datetime` DATETIME DEFAULT CURRENT_TIMESTAMP(), ' \
          '`station_number` INT(11) DEFAULT NULL, ' \
          '`station_name` VARCHAR(128) DEFAULT NULL, ' \
          '`region` VARCHAR(128) DEFAULT NULL, ' \
          '`address` VARCHAR(1024) DEFAULT NULL, ' \
          '`latitude` FLOAT DEFAULT NULL, ' \
          '`longitude` FLOAT DEFAULT NULL, ' \
          '`install_date` DATETIME DEFAULT NULL, ' \
          '`lcd_count` INT(11) DEFAULT NULL, ' \
          '`qr_count` INT(11) DEFAULT NULL, ' \
          '`proc_type` VARCHAR(128) DEFAULT NULL, ' \
          'KEY `id` (`id`)' \
          ') ENGINE=INNODB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci;'
    db.execute_and_commit(sql)

    # todo: 엑셀의 row값들을 읽어서 DB에 해당 테이블에 삽입
    # ...

    db.disconnect_db()


###################################################################################################
# Task 2
# DB에 있는 자신의계정_bicycle 테이블에서 특정 데이터를 뽑아서, 엑셀로 저장하기
# ex) 2020년 이후에 서초구에 설치된 자전거 대여소 목록데이터
# sql = 'select * from elias_bicycle where date(install_date) >= "2020-01-01" and region = "서초구";'

# sql = 'select * from elias_bicycle where date(install_date) >= %s and region = %s;'
# from_date : "2020-01-01"
# region : "서초구"
# values = (from_date, region)
# execute_and_return(sql, values)
###################################################################################################
def get_data_from_db(from_date, region, output_file_name):
    # Create new workbook
    wb = Workbook()

    # Select Worksheet
    ws = wb.active

    # Rename Worksheet
    ws.title = '대여소현황'

    # todo: Header 생성
    # ...

    # DB 객체 생성 후 연결
    db = Database(DB_URL, DB_USER, DB_PW, DB_NAME)
    db.connect_db()

    # 조건에 맞는 데이터 가져오기
    sql = 'select * from elias_bicycle where date(install_date) >= %s and region = %s;'
    values = (from_date, region)
    data_list = db.execute_and_return(sql, values)

    # todo: data_list를 가지고 엑셀의 데이터 추가
    # ...

    wb.save(output_file_name)


if __name__ == '__main__':
    put_data_to_db('public_bicycle.xlsx')
    get_data_from_db('2020-01-01', '서초구', 'new_excel.xlsx')