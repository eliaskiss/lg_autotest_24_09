from openpyxl import load_workbook
from datetime import datetime
from icecream import ic
import time

file_name = 'public_bicycle.xlsx'

ic.configureOutput(includeContext=True)

######################################################
# Open Excel File
######################################################
wb = load_workbook(file_name,
                   # read_only=True, # True: Lazy Loading 발생해서 대용량 엑셀파일 처리에 용이함
                   read_only=False, # True: Lazy Loading 발생해서 대용량 엑셀파일 처리에 용이함
                   # data_only=True) # True: Cell의 함수 결과값(3), False: Cell의 함수(=A1+A2)
                   data_only=False) # True: Cell의 함수 결과값(3), False: Cell의 함수(=A1+A2)

######################################################
# Get Sheet List
######################################################
ws_list = wb.sheetnames
# ic(ws_list)

######################################################
# Select Worksheet
######################################################
# ws = wb.active  # 현재 활성화 되어있는(저장하는순간의 워크시트) 시트를 선택
ws = wb['대여소현황']
# ws = wb[ws_list[2]]
# ws = wb[wb.sheetnames[2]]

######################################################
# Get Cell Value
######################################################
cell_a1 = ws['A1']
# ic(cell_a1)
# ic(type(cell_a1))
# ic(cell_a1.value)

######################################################
# Get Formular Cell Value
######################################################
cell_total_lcd = ws['B2593']
# ic(cell_total_lcd.value)
cell_total_lcd = ws['B2594']
# ic(cell_total_lcd.value)

######################################################
# Get Datetime Cell Value
######################################################
cell_datetime = ws['G6']
# ic(type(cell_datetime.value))
# ic(cell_datetime.value)
cell_datetime_value = cell_datetime.value.strftime('%Y-%m-%d %H:%M:%S')
# ic(cell_datetime_value)

######################################################
# Get Time Cell Value
######################################################












